import os
import subprocess
import json
from pathlib import Path
from openpyxl import load_workbook
from modify_lmp_charges import modify_lmp_charges  # 确保该模块存在并可导入
import logging

# 配置日志记录
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 路径配置，根据实际情况修改
fileGeneratePath = Path('/home/iei/share/wangyh/molyte_v1/file_wyh/')
initialSaltsPath = Path('/home/iei/share/wangyh/molyte_v1/inital_salts/')
ligpargenPath = Path('/home/iei/.conda/envs/py37/bin/')
packmolPath = Path('/home/iei/software/packmol/packmol')
lempyfyPath = Path('/home/iei/software/moltemplate/moltemplate/ltemplify.py')
moltemplatePath = Path('/home/iei/software/moltemplate/moltemplate/scripts/moltemplate.sh')
respPath = Path('/home/iei/share/wangyh/molyte_v1/RESP/')
chargeSavePath = Path('/home/iei/share/wangyh/molyte_v1/charge/')
existing_molecules_file = Path('existing_molecules.json')  # 用于存储已有溶剂信息的文件
xlsxName = '../inputList/inputList_YH3.xlsx'  # Excel文件路径请根据需要修改

# 确保必要的目录存在
fileGeneratePath.mkdir(parents=True, exist_ok=True)
chargeSavePath.mkdir(parents=True, exist_ok=True)

# 模拟参数，根据需要修改
Nsteps_NPT = 5000000
Nsteps_NVT = 15000000
Freq_trj_npt = 10000
Freq_trj_nvt = 50000
# Temp_NPT = 300
# Temp_NVT = 300
thermo_freq = 10000
timestep = 1

def run_command(command):
    """
    运行shell命令，并记录输出和错误信息。s
    """
    try:
        logging.info(f"Running command: {command}")
        result = subprocess.run(command, shell=True, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        output = result.stdout.decode().strip()
        if output:
            logging.info(output)
    except subprocess.CalledProcessError as e:
        logging.error(f"Command failed: {e.stderr.decode().strip()}")

class CurrentRow:
    """
    用于存储Excel中每一行数据的类。
    """
    pass

def get_unique_name(name, smile, existing_molecules, molecule_type):
    """
    为溶剂生成唯一名称。如果名称已存在且SMILE不同，则添加后缀生成新名称。
    """
    if name not in existing_molecules[molecule_type]:
        existing_molecules[molecule_type][name] = smile
        return name
    else:
        if existing_molecules[molecule_type][name] == smile:
            return name
        else:
            suffix = 1
            while True:
                new_name = f"{name}_{suffix}"
                if new_name not in existing_molecules[molecule_type]:
                    existing_molecules[molecule_type][new_name] = smile
                    return new_name
                suffix += 1

def find_molecule_columns(headers, prefix):
    """
    根据前缀查找所有相关的列组。

    对于 'cation' 和 'anion'：
        - 'prefix1', 'prefix1Ratio', 'prefix1Number'
        - 'prefix2', 'prefix2Ratio', 'prefix2Number'
        - ...

    对于 'sol'：
        - 'sol1', 'sol1Ratio', 'sol1Number', 'sol1SMILE'
        - 'sol2', 'sol2Ratio', 'sol2Number', 'sol2SMILE'
        - ...

    返回一个字典列表，每个字典包含对应的列名。
    """
    groups = []
    i = 1
    while True:
        name_col = f'{prefix}{i}'
        ratio_col = f'{prefix}{i}Ratio'
        number_col = f'{prefix}{i}Number'

        if prefix == 'sol':
            smile_col = f'{prefix}{i}SMILE'
            if all(col in headers for col in [name_col, ratio_col, number_col, smile_col]):
                groups.append({
                    'name_col': name_col,
                    'ratio_col': ratio_col,
                    'number_col': number_col,
                    'smile_col': smile_col
                })
            else:
                break
        else:
            if all(col in headers for col in [name_col, ratio_col, number_col]):
                groups.append({
                    'name_col': name_col,
                    'ratio_col': ratio_col,
                    'number_col': number_col
                })
            else:
                break
        i +=1
    return groups

def get_molecules(row, column_groups, include_smile=False, check_uniqueness=False, existing_molecules=None, molecule_type=None):
    """
    根据列组获取分子信息。

    参数：
    - row: 当前行的数据对象。
    - column_groups: 列组列表，每个列组是一个字典，包含列名。
    - include_smile: 是否包含SMILE信息（仅对solvent有效）。
    - check_uniqueness: 是否进行唯一性检查（仅对solvent有效）。
    - existing_molecules: 已存在的分子字典（仅对solvent有效）。
    - molecule_type: 分子类型（'solvent'），仅对solvent有效。

    返回：
    - 分子属性列表。
    """
    molecules = []
    for group in column_groups:
        name = getattr(row, group['name_col'], None)
        ratio = getattr(row, group['ratio_col'], None)
        number = getattr(row, group['number_col'], None)
        smile = getattr(row, group.get('smile_col', None), None) if include_smile else None

        if name and number:
            try:
                number_float = float(number)
                number_int = int(round(number_float))
                if number_int > 0:
                    if check_uniqueness and molecule_type and existing_molecules is not None and smile:
                        unique_name = get_unique_name(name, smile, existing_molecules, molecule_type)
                    else:
                        unique_name = name

                    molecule = {
                        'original_name': name,
                        'name': unique_name,
                        'ratio': ratio,
                        'number': number_int
                    }
                    if include_smile and smile:
                        molecule['smile'] = smile

                    molecules.append(molecule)
                    if check_uniqueness and molecule_type and existing_molecules is not None and include_smile and smile:
                        logging.info(f"Included {group['name_col']}: {unique_name}, number: {number_int} (original: {number_float}), SMILE: {smile}")
                    else:
                        logging.info(f"Included {group['name_col']}: {unique_name}, number: {number_int} (original: {number_float})")
                else:
                    logging.info(f"Skipped {group['name_col']}: {name}, number: {number_int} (original: {number_float}) <= 0")
            except ValueError:
                logging.warning(f"Invalid number '{number}' for {group['name_col']} in row '{row.name}'")
    return molecules

def elementType(element):
    """
    根据原子质量返回元素符号。
    """
    return {
        1: 'H', 4: 'He', 7: 'Li', 9: 'Be', 11: 'B', 12: 'C',
        14: 'N', 16: 'O', 19: 'F', 20: 'Ne', 23: 'Na', 24: 'Mg',
        27: 'Al', 28: 'Si', 31: 'P', 32: 'S', 35: 'Cl', 39: 'K',
        40: 'Ca', 45: 'Sc', 48: 'Ti', 51: 'V', 52: 'Cr', 55: 'Mn',
        56: 'Fe', 59: 'Co', 64: 'Cu', 65: 'Zn', 70: 'Ga', 73: 'Ge',
        75: 'As', 79: 'Se'
    }.get(element, 'A')

# 从文件中加载已有溶剂信息（existing_molecules）
if existing_molecules_file.exists():
    with open(existing_molecules_file, 'r') as f:
        existing_molecules = json.load(f)
    if 'solvent' not in existing_molecules:
        existing_molecules['solvent'] = {}
else:
    existing_molecules = {'solvent': {}}

# 加载Excel文件
wb = load_workbook(xlsxName, data_only=True)
sheet1 = wb.active
totalRows = sheet1.max_row
logging.info(f"Total Rows: {totalRows}")

# 读取表头（假设表头在第二行）
headers = [cell.value for cell in sheet1[2]]
logging.info(f"Headers: {headers}")

# 查找分子列组
cation_groups = find_molecule_columns(headers, 'cation')
anion_groups = find_molecule_columns(headers, 'anion')
solvent_groups = find_molecule_columns(headers, 'sol')

logging.info(f"Detected {len(cation_groups)} cation(s).")
logging.info(f"Detected {len(anion_groups)} anion(s).")
logging.info(f"Detected {len(solvent_groups)} solvent(s).")

for row_idx in range(3, totalRows + 1):  # 从第3行开始，假设前两行是标题
    currentRow = CurrentRow()
    for cell in sheet1[row_idx]:
        header = sheet1.cell(row=2, column=cell.column).value
        setattr(currentRow, header, cell.value)

    # 处理阳离子
    cations = get_molecules(
        row=currentRow,
        column_groups=cation_groups,
        include_smile=False,
        check_uniqueness=False
    )

    # 处理阴离子
    anions = get_molecules(
        row=currentRow,
        column_groups=anion_groups,
        include_smile=False,
        check_uniqueness=False
    )

    # 处理溶剂
    solvents = get_molecules(
        row=currentRow,
        column_groups=solvent_groups,
        include_smile=True,
        check_uniqueness=True,
        existing_molecules=existing_molecules,
        molecule_type='solvent'
    )

    logging.info(f"\nProcessing Row {row_idx}: {currentRow.name}")
    logging.info(f"Cations: {cations}")
    logging.info(f"Anions: {anions}")
    logging.info(f"Solvents: {solvents}")

    # 至少有一个阳离子数量 > 0 才继续
    valid_cation = any(cat['number'] > 0 for cat in cations)
    if not valid_cation:
        logging.warning(f"Skipping row {row_idx} due to no valid cation.")
        continue

    # 创建目标目录
    destination_path = fileGeneratePath / currentRow.name
    if destination_path.exists():
        logging.info(f"Directory {destination_path} already exists. Removing it.")
        subprocess.run(f'rm -rf {destination_path}', shell=True)
    run_command(f'cp -rf {initialSaltsPath} {fileGeneratePath}')
    run_command(f'mv -f {fileGeneratePath}/inital_salts/ {destination_path}')
    os.chdir(destination_path)
    current_path = os.getcwd()

    # 生成 SLURM 作业脚本（job.sh）
    job_script_content = f"""#!/bin/bash
#PBS -N {currentRow.name}
#PBS -o out.dat
#PBS -e err.dat
#PBS -q batch
#PBS -l nodes=1:ppn=64
#PBS -l walltime=7200:00:00
cd $PBS_O_WORKDIR

source /home/iei/share/software/environment/lammps.env
EXEC=/home/iei/share/software/lammps/200303/lmp_mpi
input=all.in
mpirun -n 64 $EXEC <{currentRow.name}.in >{currentRow.name}.log
"""

    with open(destination_path / 'job.sh', 'w') as job_file:
        job_file.write(job_script_content)
        logging.info(f"Generated job.sh for {currentRow.name}")

    # 处理溶剂：生成电荷文件
    for solvent in solvents:
        if solvent['name']:
            chargeFilePath = chargeSavePath / f"{solvent['name']}.charmm.chg"
            # 调用 ligpargen 生成初始电荷文件
            run_command(f"{ligpargenPath}/ligpargen -s '{solvent['smile']}' -n {solvent['name']} -r MOL -c 0 -o 0 -cgen CM1A")
            if not chargeFilePath.exists():
                logging.info(f"Generating charges for {solvent['name']}")
                run_command(f"{respPath}/RESP2.sh '{solvent['name']}.charmm.pdb'")
                run_command(f"cp {solvent['name']}.charmm.chg {chargeFilePath}")
            else:
                logging.info(f"Charge file for {solvent['name']} already exists. Skipping generation.")
            # 修改 LAMMPS 文件中的电荷
            modify_lmp_charges(f"{solvent['name']}.lammps.lmp", str(chargeFilePath), f"{solvent['name']}.lmp")

    # 生成 Packmol 输入文件
    packmol_input = f"tolerance 2.0\nfiletype pdb\noutput {currentRow.name}.pdb\n"
    # 加入阳离子
    for cat in cations:
        if cat['name'] and cat['number'] > 0:
            packmol_input += f"""structure {cat['name']}.pdb
   number {cat['number']}
   inside box 0. 0. 0 {currentRow.V} {currentRow.V} {currentRow.V}
end structure
"""
    # 加入阴离子
    for anion in anions:
        if anion['name'] and anion['number'] > 0:
            packmol_input += f"""structure {anion['name']}.pdb
   number {anion['number']}
   inside box 0. 0. 0 {currentRow.V} {currentRow.V} {currentRow.V}
end structure
"""
    # 加入溶剂
    for solvent in solvents:
        if solvent['name'] and solvent['number'] > 0:
            packmol_input += f"""structure {solvent['name']}.charmm.pdb
   number {solvent['number']}
   inside box 0. 0. 0 {currentRow.V} {currentRow.V} {currentRow.V}
end structure
"""

    packmol_inp_path = destination_path / f"{currentRow.name}.inp"
    with open(packmol_inp_path, 'w') as packmol_file:
        packmol_file.write(packmol_input)
        logging.info(f"Generated Packmol input file for {currentRow.name}")

    # 运行 Packmol
    run_command(f"{packmolPath} < {packmol_inp_path}")

    # 为溶剂生成 Moltemplate LT 文件
    for solvent in solvents:
        if solvent['name']:
            run_command(f"{lempyfyPath} -name {solvent['name']} {destination_path}/{solvent['name']}.lmp > {destination_path}/{solvent['name']}.lt")

    # 生成 Moltemplate LT 文件
    moltemplate_content = ""
    # 导入阳离子、阴离子和溶剂的 LT 文件
    for cat in cations:
        if cat['name'] and cat['number'] > 0:
            moltemplate_content += f'import "{cat["name"]}.lt"\n'
    for anion in anions:
        if anion['name'] and anion['number'] > 0:
            moltemplate_content += f'import "{anion["name"]}.lt"\n'
    for solvent in solvents:
        if solvent['name'] and solvent['number'] > 0:
            moltemplate_content += f'import "{solvent["name"]}.lt"\n'

    # 实例化分子
    for cat in cations:
        if cat['name'] and cat['number'] > 0:
            moltemplate_content += f'{cat["name"]}s = new {cat["name"]}[{cat["number"]}]\n'
    for anion in anions:
        if anion['name'] and anion['number'] > 0:
            moltemplate_content += f'{anion["name"]}s = new {anion["name"]}[{anion["number"]}]\n'
    for solvent in solvents:
        if solvent['name'] and solvent['number'] > 0:
            moltemplate_content += f'{solvent["name"]}s = new {solvent["name"]}[{solvent["number"]}]\n'

    # 使用双大括号转义，以在输出中生成单个大括号
    moltemplate_content += f"""
write_once("Data Boundary") {{
    0 {currentRow.V} xlo xhi
    0 {currentRow.V} ylo yhi
    0 {currentRow.V} zlo zhi
}}
"""

    moltemplate_lt_path = destination_path / f"{currentRow.name}.lt"
    with open(moltemplate_lt_path, 'w') as moltemplate_file:
        moltemplate_file.write(moltemplate_content)
        logging.info(f"Generated Moltemplate LT file for {currentRow.name}")

    # 运行 Moltemplate 脚本生成 LAMMPS 输入文件
    run_command(f"{moltemplatePath} -pdb {currentRow.name}.pdb {moltemplate_lt_path}")

    # 检查 LAMMPS 数据文件是否存在
    lmp_data_path = destination_path / f"{currentRow.name}.data"
    if not lmp_data_path.exists():
        logging.error(f"LAMMPS data file {lmp_data_path} does not exist.")
        continue

    # 解析 LAMMPS 数据文件中的 Masses 部分
    with open(lmp_data_path, 'r', encoding='utf-8') as lmpDataModify:
        contentList = lmpDataModify.readlines()

    lmpListModifyArr = []
    for i, line in enumerate(contentList):
        if line.strip().lower() == 'masses':
            for j in range(i + 2, len(contentList)):
                if not contentList[j].strip() or 'atoms' in contentList[j].lower():
                    break
                temp = contentList[j].split()
                if len(temp) < 2:
                    logging.warning(f"Malformed Masses line: '{contentList[j]}'")
                    continue
                try:
                    atomic_mass = round(float(temp[1]))
                    element = elementType(atomic_mass)
                    lmpListModifyArr.append(element)
                except ValueError:
                    logging.warning(f"Invalid atomic mass '{temp[1]}' in Masses line.")
            break

    findLiO = ''
    fixRdf = ''
    # 根据需要，如果需要对特定元素对进行 RDF 分析，可在此添加逻辑
    # 例如，如果需要 Li-O 对，可以添加以下逻辑
    # 这里只是示例，具体逻辑根据实际需求调整
    # 定义阳离子元素，可以根据需要修改或增加更多的阳离子
    cation_elements = ['Li', 'Na', 'K', 'Mg', 'Ca','Zn','Al']

    # 遍历 lmpListModifyArr 列表中的每个元素
    for idx, value in enumerate(lmpListModifyArr, start=1):
    # 如果该元素是阳离子元素
        if value in cation_elements:
        # 继续对每个阳离子元素和目标阴离子元素生成 RDF 配对
            for anion_idx, anion in enumerate(lmpListModifyArr, start=1):
            # 阴离子元素列表可以根据实际需求进行修改
                if anion in ['O', 'N', 'P', 'B']:  # 这里只是一个示例，阴离子列表可以调整
                # 为阳离子和阴离子之间的配对生成 RDF 配对信息
                    findLiO += f'{idx} {anion_idx} '  # idx 是阳离子元素的索引，anion_idx 是阴离子元素的索引
                    fixRdf += f'rdf_{anion} '  # 为每个阴离子元素生成 RDF 标签

#    for idx, value in enumerate(lmpListModifyArr, start=1):
#        if value in ['O', 'P', 'B', 'N']:
#            findLiO += f'1 {idx} '
#            fixRdf += f'rdf_{value} '

    in_list_content = f'variable element_list index "{" ".join(lmpListModifyArr)}"\n'
    in_list_content += f'variable rdf_pair string "{findLiO.strip()}"\n'

    in_list_path = fileGeneratePath / currentRow.name / f"{currentRow.name}.in.list"
    with open(in_list_path, 'w', encoding='utf-8') as lmpList:
        lmpList.write(in_list_content)
        logging.info(f"Generated LAMMPS .in.list file for {currentRow.name}")

    # 生成 LAMMPS 输入文件（.in）
    lmp_in_content = f'''# ----------------- Variable Section -----------------
'''
    # 定义变量：Cation, Anion, Solvent
    for i, cat in enumerate(cations, start=1):
        lmp_in_content += f'variable Cation{i} index {cat["name"]}\n'
    for i, anion in enumerate(anions, start=1):
        lmp_in_content += f'variable Anion{i} index {anion["name"]}\n'
    for i, solvent in enumerate(solvents, start=1):
        lmp_in_content += f'variable Solvent{i} index {solvent["name"]}\n'

    # 定义其他变量
    lmp_in_content += f'''
variable infile string {currentRow.name}
variable outname string {currentRow.name}

include {currentRow.name}.in.init
read_data {currentRow.name}.data
include {currentRow.name}.in.settings
include {currentRow.name}.in.list
include {currentRow.name}.in.list_salt

variable Nsteps_NPT equal {Nsteps_NPT}
variable Nsteps_NVT equal {Nsteps_NVT}

variable Freq_trj_npt index {Freq_trj_npt}
variable Freq_trj_nvt index {Freq_trj_nvt}

variable Temp_NPT equal {currentRow.temp}
variable Temp_NVT equal {currentRow.temp}

variable thermo_freq equal {thermo_freq}
variable timestep equal {timestep}
'''

    # 定义 group，根据 cations, anions, solvents 创建相应的 group
    for i, cat in enumerate(cations, start=1):
        if cat['name'] and cat['number'] > 0:
            cat_group = f'Cation{i}'
            lmp_in_content += f'group {cat_group} union {cat["name"]}\n'

    for i, anion in enumerate(anions, start=1):
        if anion['name'] and anion['number'] > 0:
            anion_group = f'Anion{i}'
            lmp_in_content += f'group {anion_group} union {anion["name"]}\n'

#    for i, solvent in enumerate(solvents, start=1):
#        if solvent['name'] and solvent['number'] > 0:
#            solvent_group = f'Solvent{i}'
#            lmp_in_content += f'group {solvent_group} union {solvent["name"]}\n'

    # 使用双大括号转义，以在输出中生成单个大括号
    lmp_in_content += f'''
thermo_style custom step cpu cpuremain temp density lx ly lz etotal ke pe evdwl ecoul elong ebond eangle edihed eimp
thermo ${{thermo_freq}}
timestep ${{timestep}}

minimize 1.0e-4 1.0e-6 5000 10000

write_data ${{infile}}_after_minimize.data nocoeff
write_dump all custom ${{infile}}_after_minimize.lammpstrj id element mol type x y z q modify element ${{element_list}} sort id

reset_timestep 0

dump trj_npt all custom ${{Freq_trj_npt}} NPT_${{outname}}.lammpstrj id element mol type x y z q
dump_modify trj_npt flush yes element ${{element_list}} sort id

fix fxnpt all npt temp ${{currentRow.temp}} ${{currentRow.temp}} $(100.0*dt) iso 0.0 0.0 $(1000.0*dt)
run ${{Nsteps_NPT}}
unfix fxnpt

undump trj_npt

write_data ${{infile}}_after_npt.data
write_restart ${{infile}}_restart_after_npt.data
write_dump all custom ${{infile}}_after_npt.lammpstrj id element mol type x y z q modify element ${{element_list}} sort id

reset_timestep 0
'''

    # 对每个阴离子进行后处理（MSD 计算）
    for i, anion in enumerate(anions, start=1):
        if anion['name'] and anion['number'] > 0:
            anion_group = anion['name']
            lmp_in_content += f'compute An{i} {anion_group} msd com yes\n'
            lmp_in_content += f'fix An{i}msd {anion_group} ave/time {Freq_trj_nvt} 1 {Freq_trj_nvt} c_An{i}[1] c_An{i}[2] c_An{i}[3] c_An{i}[4] file out_{anion["name"]}_msd.dat title1 "t msd msd msd msd_{anion["name"]}" title2 "fs {anion["name"]}_x {anion["name"]}_y {anion["name"]}_z {anion["name"]}_total"\n'

    # 对每个阳离子进行后处理（MSD 计算）
    for i, cation in enumerate(cations, start=1):
        if cation['name'] and cation['number'] > 0:
            cation_group = cation['name']
            lmp_in_content += f'compute Ca{i} {cation_group} msd com yes\n'
            lmp_in_content += f'fix Ca{i}msd {cation_group} ave/time {Freq_trj_nvt} 1 {Freq_trj_nvt} c_Ca{i}[1] c_Ca{i}[2] c_Ca{i}[3] c_Ca{i}[4] file out_{cation["name"]}_msd.dat title1 "t msd msd msd msd_{cation["name"]}" title2 "fs {cation["name"]}_x {cation["name"]}_y {cation["name"]}_z {cation["name"]}_total"\n'

    # RDF 计算相关设置，转义花括号
    lmp_in_content += f'''
compute rdfc1 all rdf 100 ${{rdf_pair}}
fix rdff1 all ave/time $(v_Nsteps_NVT/1000) 1000 ${{Nsteps_NVT}} c_rdfc1[*] file out_rdf.dat mode vector title3 "RDF {fixRdf}"

dump trj_nvt all custom ${{Freq_trj_nvt}} NVT_${{outname}}.lammpstrj id element mol type x y z q
dump_modify trj_nvt flush yes element ${{element_list}} sort id
dump utrj_nvt all custom ${{Freq_trj_nvt}} NVT_${{outname}}_un.lammpstrj id element mol type xu yu zu ix iy iz q
dump_modify utrj_nvt flush yes element ${{element_list}} sort id

fix fxnvt all nvt temp ${{currentRow.temp}} ${{currentRow.temp}} $(100.0*dt)
run ${{Nsteps_NVT}}
unfix fxnvt

undump trj_nvt
undump utrj_nvt

write_data ${{infile}}_after_nvt.data
write_restart ${{infile}}_restart_after_nvt.data
write_dump all custom ${{infile}}_after_nvt.lammpstrj id element mol type x y z q modify element ${{element_list}} sort id
'''

    # 写入 LAMMPS 输入文件
    with open(destination_path / f"{currentRow.name}.in", 'w') as lmpIn:
        lmpIn.write(lmp_in_content)
        logging.info(f"Generated LAMMPS .in file for {currentRow.name}")

    # 提交 SLURM 作业
#    run_command(f'qsub {destination_path}/job.sh')

# 脚本结束后保存更新后的 existing_molecules（仅溶剂信息）
with open(existing_molecules_file, 'w') as f:
    json.dump(existing_molecules, f, indent=4)
logging.info("Updated existing_molecules.json with the latest molecule information.")

