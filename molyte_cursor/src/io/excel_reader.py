"""
Excel读取器模块，负责从Excel文件中读取和解析数据
"""
import json
import re
import pandas as pd
import hashlib
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Union, Any

from ..utils.logger import Logger
from ..models.molecule import Cation, Anion, Solvent, MoleculeSystem
from ..utils.molecule_calculator import MoleculeCalculator
from .excel_processor import ExcelProcessor

class ExcelReader:
    """Excel文件读取器类"""
    
    # 支持的阳离子和阴离子列表
    SUPPORTED_CATIONS = ['Li', 'Na', 'K', 'Zn', 'Mg', 'Mg0.8', 'Zn0.8']
    SUPPORTED_ANIONS = ['PF6', 'FSI', 'TFSI', 'DFOB', 'Cl', 'NO3', 'ClO4', 'BF4']
    
    def __init__(self, existing_molecules_file=None, calculation_history_file=None, auto_calculate=True):
        """初始化Excel读取器
        
        Args:
            existing_molecules_file: 现有分子信息文件路径
            calculation_history_file: 计算历史记录文件路径
            auto_calculate: 是否自动计算分子数量（而不是从Excel读取）
        """
        self.logger = Logger().get_logger()
        self.existing_molecules_file = Path(existing_molecules_file) if existing_molecules_file else None
        self.calculation_history_file = Path(calculation_history_file) if calculation_history_file else Path("calculation_history.json")
        self.existing_molecules = self._load_existing_molecules()
        self.calculation_history = self._load_calculation_history()
        self.auto_calculate = auto_calculate
        self.rejected_components = []  # 存储被拒绝的组分和原因
        
        # 初始化分子计算器和Excel处理器
        self.molecule_calculator = MoleculeCalculator(logger=self.logger)
        self.excel_processor = ExcelProcessor(logger=self.logger)
    
    def _load_existing_molecules(self):
        """加载现有分子信息
        
        Returns:
            现有分子字典
        """
        if self.existing_molecules_file and self.existing_molecules_file.exists():
            with open(self.existing_molecules_file, 'r') as f:
                try:
                    return json.load(f)
                except json.JSONDecodeError:
                    self.logger.error(f"无法解析分子信息文件: {self.existing_molecules_file}")
                    return {}
        return {}
    
    def _load_calculation_history(self):
        """加载计算历史记录
        
        Returns:
            计算历史字典
        """
        if self.calculation_history_file and self.calculation_history_file.exists():
            with open(self.calculation_history_file, 'r') as f:
                try:
                    return json.load(f)
                except json.JSONDecodeError:
                    self.logger.error(f"无法解析计算历史文件: {self.calculation_history_file}")
                    return {}
        return {}
    
    def save_existing_molecules(self):
        """保存现有分子信息"""
        if self.existing_molecules_file:
            with open(self.existing_molecules_file, 'w') as f:
                json.dump(self.existing_molecules, f, indent=2)
                self.logger.info(f"分子信息已保存到: {self.existing_molecules_file}")
    
    def save_calculation_history(self):
        """保存计算历史记录"""
        with open(self.calculation_history_file, 'w') as f:
            json.dump(self.calculation_history, f, indent=2)
            self.logger.info(f"计算历史已保存到: {self.calculation_history_file}")
    
    def save_rejected_components(self):
        """保存被拒绝的组分列表"""
        with open("rejected_components.txt", "w") as f:
            f.write("拒绝的组分列表和原因：\n")
            for item in self.rejected_components:
                f.write(f"{item['system_name']} - {item['component_type']}: {item['component_name']} - {item['reason']}\n")
        self.logger.info("已保存拒绝组分列表")
    
    def find_molecule_columns(self, headers, prefix):
        """查找特定前缀的分子列
        
        Args:
            headers: 表头列表
            prefix: 列名前缀
            
        Returns:
            匹配的列名和对应索引的字典
        """
        columns = {}
        pattern = re.compile(f"{prefix}(\\d+)", re.IGNORECASE)
        
        for i, header in enumerate(headers):
            if isinstance(header, str):
                match = pattern.match(header)
                if match:
                    index = int(match.group(1))
                    columns[header] = index
        
        # 按索引排序
        sorted_columns = dict(sorted(columns.items(), key=lambda item: item[1]))
        self.logger.debug(f"找到{prefix}列: {sorted_columns}")
        return sorted_columns
    
    def is_supported_ion(self, name, ion_type):
        """检查离子是否被支持
        
        Args:
            name: 离子名称
            ion_type: 离子类型('cation'或'anion')
            
        Returns:
            布尔值，True表示支持
        """
        if ion_type == 'cation':
            return name in self.SUPPORTED_CATIONS
        elif ion_type == 'anion':
            return name in self.SUPPORTED_ANIONS
        return True  # 溶剂默认支持
    
    def get_unique_name(self, name, smile, molecule_type):
        """生成唯一的分子名称
        
        Args:
            name: 分子名称
            smile: SMILE字符串
            molecule_type: 分子类型
            
        Returns:
            唯一的分子名称
        """
        # 检查name是否有效
        if not name or pd.isna(name):
            name = f"Unknown_{molecule_type}"
        
        # 检查该名称是否已存在
        if molecule_type in self.existing_molecules:
            for existing_name, existing_smile in self.existing_molecules[molecule_type].items():
                # 如果SMILE相同，返回已有名称
                if existing_smile == smile:
                    self.logger.debug(f"已找到相同的{molecule_type}: {existing_name}")
                    return existing_name
            
            # 如果名称已存在但SMILE不同，生成新名称
            if name in self.existing_molecules[molecule_type]:
                counter = 1
                while f"{name}_{counter}" in self.existing_molecules[molecule_type]:
                    counter += 1
                name = f"{name}_{counter}"
        else:
            self.existing_molecules[molecule_type] = {}
        
        # 添加到现有分子中
        if molecule_type == "solvent" and smile:  # 只有溶剂需要存储SMILE
            self.existing_molecules[molecule_type][name] = smile
        else:
            self.existing_molecules[molecule_type][name] = ""
        return name
    
    def get_molecules(self, row_data, column_groups, include_smile=False, check_uniqueness=False, molecule_type=None, system_name="Unknown"):
        """从行数据中提取分子信息
        
        Args:
            row_data: Excel行数据
            column_groups: 列分组信息
            include_smile: 是否包含SMILE字符串
            check_uniqueness: 是否检查唯一性
            molecule_type: 分子类型
            system_name: 系统名称，用于记录拒绝原因
            
        Returns:
            分子列表
        """
        molecules = []
        
        for col_name, col_idx in column_groups.items():
            # 获取分子数量
            count = row_data.get(col_name, 0)
            if pd.isna(count) or count == 0:
                continue
            
            # 获取分子名称和SMILE
            name_col = f"{col_name}_name"
            smile_col = f"{col_name}_smile"
            ratio_col = f"{col_name}_ratio"
            
            name = row_data.get(name_col, col_name)
            smile = row_data.get(smile_col, "") if molecule_type == "solvent" else ""  # 只有溶剂需要SMILE
            ratio = row_data.get(ratio_col, None)
            
            # 检查阴阳离子是否被支持
            if (molecule_type == 'cation' or molecule_type == 'anion') and not self.is_supported_ion(name, molecule_type):
                self.logger.warning(f"不支持的{molecule_type}: {name}，将被记录但继续处理其他组分")
                self.rejected_components.append({
                    'system_name': system_name,
                    'component_type': molecule_type,
                    'component_name': name,
                    'reason': f"不支持的{molecule_type}类型"
                })
                continue
            
            # 如果需要检查唯一性
            if check_uniqueness:
                name = self.get_unique_name(name, smile, molecule_type)
            
            # 创建分子对象
            molecule_data = {
                "name": name,
                "count": int(count)
            }
            
            if include_smile and smile and molecule_type == "solvent":
                molecule_data["smile"] = smile
                
            if ratio is not None and not pd.isna(ratio):
                molecule_data["ratio"] = float(ratio)
            
            molecules.append(molecule_data)
        
        return molecules
    
    def extract_simulation_params(self, row_data, system_name="Unknown"):
        """提取模拟参数
        
        Args:
            row_data: Excel行数据
            system_name: 系统名称，用于记录拒绝原因
            
        Returns:
            模拟参数字典
        """
        params = {}
        
        # 提取常见的模拟参数
        param_keys = [
            'box_size_x', 'box_size_y', 'box_size_z',  # 盒子尺寸
            'box_size',                                # 立方盒子尺寸
            'concentration', 'temperature',             # 浓度和温度
            'solvent_ratio', 'cation_anion_ratio',      # 比例参数
            'density', 'cutoff'                         # 其他参数
        ]
        
        for key in param_keys:
            if key in row_data and not pd.isna(row_data[key]):
                params[key] = row_data[key]
        
        # 处理盒子尺寸
        if all(f'box_size_{dim}' in params for dim in ['x', 'y', 'z']):
            params['box_size'] = (
                params['box_size_x'],
                params['box_size_y'],
                params['box_size_z']
            )
        elif 'box_size' not in params:
            # 设置默认盒子尺寸为40埃
            params['box_size'] = 40.0
            self.logger.warning("未指定盒子尺寸，使用默认值: 40.0 埃")
        
        # 验证盒子尺寸范围
        if 'box_size' in params:
            box_size = params['box_size']
            if isinstance(box_size, (int, float)) and (box_size <= 0 or box_size > 100):
                self.logger.warning(f"盒子尺寸 {box_size} 不在有效范围内(0-100埃)，将被记录但继续处理")
                self.rejected_components.append({
                    'system_name': system_name,
                    'component_type': 'system_parameter',
                    'component_name': 'box_size',
                    'reason': f"无效的盒子尺寸: {box_size}，必须在0-100范围内"
                })
            
        # 确保温度参数存在
        if 'temperature' not in params:
            params['temperature'] = 300.0  # 默认温度为300K
            self.logger.warning("未指定温度，使用默认值: 300.0 K")
        
        return params
    
    def generate_system_hash(self, system_data):
        """生成系统唯一哈希值
        
        Args:
            system_data: 系统数据
            
        Returns:
            哈希字符串
        """
        # 提取用于哈希的关键数据
        hash_data = {
            'temperature': system_data.get('temperature', 300.0),
            'components': []
        }
        
        # 添加所有组分信息
        for component_type in ['cations', 'anions', 'solvents']:
            if component_type in system_data:
                for comp in system_data[component_type]:
                    hash_data['components'].append({
                        'name': comp.get('name', ''),
                        'smile': comp.get('smile', ''),
                        'ratio': comp.get('ratio', 1.0)
                    })
        
        # 生成哈希值
        hash_str = json.dumps(hash_data, sort_keys=True)
        return hashlib.md5(hash_str.encode()).hexdigest()
    
    def check_duplicate_calculation(self, system_data):
        """检查是否已经计算过相同的系统
        
        Args:
            system_data: 系统数据
            
        Returns:
            (是否是重复计算, 先前计算的路径)
        """
        system_hash = self.generate_system_hash(system_data)
        
        if system_hash in self.calculation_history:
            previous_path = self.calculation_history[system_hash]
            self.logger.warning(f"系统已经计算过，路径: {previous_path}")
            return True, previous_path
        
        return False, None
    
    def auto_calculate_counts(self, row_data, column_groups):
        """自动计算分子数量
        
        Args:
            row_data: Excel行数据
            column_groups: 列分组信息
            
        Returns:
            更新后的行数据
        """
        # 如果自动计算功能未启用，直接返回原始数据
        if not self.auto_calculate:
            return row_data
        
        # 检查是否已经提供了分子数量
        has_valid_counts = False
        for category, columns in column_groups.items():
            for col_name in columns:
                count = row_data.get(col_name, 0)
                if not pd.isna(count) and count > 0:
                    has_valid_counts = True
                    break
            if has_valid_counts:
                break
        
        # 如果已经提供了有效的分子数量，使用这些数量
        if has_valid_counts:
            self.logger.info("使用Excel中提供的分子数量")
            return row_data
        
        # 提取模拟参数
        params = self.extract_simulation_params(row_data)
        
        # 检查是否有足够的参数进行计算
        if 'box_size' in params and 'concentration' in params:
            self.logger.info("自动计算分子数量")
            
            # 检查是否使用基于比例的计算
            use_ratio_based = False
            for col in row_data:
                if col.endswith('_ratio') and not pd.isna(row_data[col]):
                    use_ratio_based = True
                    break
            
            if use_ratio_based:
                # 基于比例的计算
                self.logger.info("使用基于比例的计算方式")
                
                # 收集组分比例
                component_ratios = {}
                for col in row_data:
                    if col.endswith('_ratio') and not pd.isna(row_data[col]):
                        component_ratios[col.replace('_ratio', '')] = float(row_data[col])
                
                # 计算分子数量
                result = self.molecule_calculator.calculate_from_primary_cation(
                    box_size=params['box_size'],
                    primary_cation_concentration=params['concentration'],
                    component_ratios=component_ratios
                )
                
                # 更新行数据
                new_row_data = row_data.copy()
                
                # 更新阳离子数量
                for component, count in result['cations'].items():
                    col_idx = int(component.replace('cation', ''))
                    col_name = f"cation{col_idx}"
                    if col_name in column_groups.get('cations', {}):
                        new_row_data[col_name] = count
                
                # 更新阴离子数量
                for component, count in result['anions'].items():
                    col_idx = int(component.replace('anion', ''))
                    col_name = f"anion{col_idx}"
                    if col_name in column_groups.get('anions', {}):
                        new_row_data[col_name] = count
                
                # 更新溶剂数量
                for component, count in result['solvents'].items():
                    col_idx = int(component.replace('sol', ''))
                    col_name = f"sol{col_idx}"
                    if col_name in column_groups.get('solvents', {}):
                        new_row_data[col_name] = count
                
                self.logger.info(f"基于比例的计算完成")
                return new_row_data
            else:
                # 传统计算方式
                # 准备计算所需的参数
                box_size = params['box_size']
                concentration = params['concentration']
                solvent_ratio = params.get('solvent_ratio', 0.0)
                cation_anion_ratio = params.get('cation_anion_ratio', 1.0)
                
                # 获取各类型的列名
                cation_cols = list(column_groups.get('cations', {}).keys())
                anion_cols = list(column_groups.get('anions', {}).keys())
                solvent_cols = list(column_groups.get('solvents', {}).keys())
                
                # 准备溶剂数据
                if solvent_ratio == 0.0 and solvent_cols:
                    # 如果没有提供溶剂比例但有溶剂列，默认使用一个标准比例
                    solvent_ratio = 10.0
                    self.logger.warning(f"未提供溶剂比例，使用默认值: {solvent_ratio}")
                
                solvent_data = {}
                for sol_col in solvent_cols:
                    # 如果Excel中没有提供具体比例，假设所有溶剂比例相等
                    solvent_data[sol_col] = 1.0 / len(solvent_cols) if solvent_cols else 0.0
                
                # 计算分子数量
                result = self.molecule_calculator.calculate_from_system_definition(
                    box_size=box_size,
                    cation_concentration=concentration,
                    solvent_data=solvent_data,
                    cation_types=len(cation_cols),
                    anion_types=len(anion_cols),
                    length_unit='angstrom',
                    concentration_unit='mol/L'
                )
                
                # 更新行数据
                new_row_data = row_data.copy()
                
                # 更新阳离子数量
                for i, col in enumerate(cation_cols):
                    key = f'cation{i+1}'
                    if key in result['cations']:
                        new_row_data[col] = result['cations'][key]
                
                # 更新阴离子数量
                for i, col in enumerate(anion_cols):
                    key = f'anion{i+1}'
                    if key in result['anions']:
                        new_row_data[col] = result['anions'][key]
                
                # 更新溶剂数量
                if solvent_data:
                    for i, col in enumerate(solvent_cols):
                        solvent_key = list(result['solvents'].keys())[i] if i < len(result['solvents']) else None
                        if solvent_key:
                            new_row_data[col] = result['solvents'][solvent_key]
                
                self.logger.info(f"自动计算的分子数量已更新")
                return new_row_data
        else:
            self.logger.warning("无法自动计算分子数量，缺少必要参数")
            return row_data
    
    def read_excel(self, excel_file):
        """读取Excel文件并解析为分子系统
        
        Args:
            excel_file: Excel文件路径
            
        Returns:
            解析后的分子系统列表
        """
        self.logger.info(f"读取Excel文件: {excel_file}")
        
        try:
            # 加载工作簿
            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            # 获取表头
            headers = [cell.value for cell in sheet[2] if cell.value is not None]
            self.logger.debug(f"表头: {headers}")
            
            # 查找分子列
            cation_columns = self.find_molecule_columns(headers, "cation")
            anion_columns = self.find_molecule_columns(headers, "anion")
            solvent_columns = self.find_molecule_columns(headers, "sol")
            
            column_groups = {
                'cations': cation_columns,
                'anions': anion_columns,
                'solvents': solvent_columns
            }
            
            # 解析数据行
            systems = []
            for row_idx in range(3, sheet.max_row + 1):
                row_data = {}
                
                # 提取单元格数据
                for col_idx, header in enumerate(headers, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    row_data[header] = cell.value
                
                # 检查是否为空行
                if not any(value is not None and value != "" for value in row_data.values()):
                    continue
                
                # 获取系统名称
                name = row_data.get('name', f"System_{row_idx}")
                
                # 如果需要自动计算分子数量
                row_data = self.auto_calculate_counts(row_data, column_groups)
                
                # 提取模拟参数
                params = self.extract_simulation_params(row_data, name)
                temperature = params.get('temperature', 300.0)
                box_size = params.get('box_size', 40.0)
                
                # 提取分子信息
                cations = self.get_molecules(row_data, cation_columns, include_smile=False, check_uniqueness=True, molecule_type='cation', system_name=name)
                anions = self.get_molecules(row_data, anion_columns, include_smile=False, check_uniqueness=True, molecule_type='anion', system_name=name)
                solvents = self.get_molecules(row_data, solvent_columns, include_smile=True, check_uniqueness=True, molecule_type='solvent', system_name=name)
                
                # 检查是否至少有一个有效的阳离子和阴离子
                if not cations:
                    self.logger.warning(f"系统 {name} 没有有效的阳离子，跳过")
                    self.rejected_components.append({
                        'system_name': name,
                        'component_type': 'system',
                        'component_name': name,
                        'reason': "没有有效的阳离子"
                    })
                    continue
                
                if not anions:
                    self.logger.warning(f"系统 {name} 没有有效的阴离子，跳过")
                    self.rejected_components.append({
                        'system_name': name,
                        'component_type': 'system',
                        'component_name': name,
                        'reason': "没有有效的阴离子"
                    })
                    continue
                
                # 准备系统数据用于检查重复计算
                system_data = {
                    'name': name,
                    'temperature': temperature,
                    'cations': cations,
                    'anions': anions,
                    'solvents': solvents
                }
                
                # 检查是否为重复计算
                is_duplicate, previous_path = self.check_duplicate_calculation(system_data)
                if is_duplicate:
                    self.logger.warning(f"系统 {name} 已经计算过，路径: {previous_path}")
                    
                    # 将重复计算信息添加到拒绝组分中
                    self.rejected_components.append({
                        'system_name': name,
                        'component_type': 'system',
                        'component_name': name,
                        'reason': f"重复计算，已有结果在: {previous_path}"
                    })
                    
                    continue
                
                # 创建分子系统
                system = MoleculeSystem(
                    name=name,
                    cations=[Cation(**cation) for cation in cations],
                    anions=[Anion(**anion) for anion in anions],
                    solvents=[Solvent(**solvent) for solvent in solvents],
                    temperature=temperature,
                    box_size=box_size
                )
                
                systems.append(system)
                self.logger.info(f"解析系统: {name}, 温度: {temperature}K, 阳离子: {len(cations)}, 阴离子: {len(anions)}, 溶剂: {len(solvents)}")
            
            # 保存分子信息和拒绝组分信息
            self.save_existing_molecules()
            self.save_rejected_components()
            
            return systems
            
        except Exception as e:
            self.logger.error(f"读取Excel文件出错: {e}")
            raise 